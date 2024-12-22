from tkinter import *
from tkinter import ttk
from PIL import Image, ImageTk
from customtkinter import *
import os
import sqlite3
# import tkinter as tk
from tkinter.messagebox import askyesno
# from datetime import date, datetime
from tkinter import messagebox #filedialog 
import openpyxl #xlrd
from openpyxl import Workbook
import pathlib
from pathlib import *
import datetime
from tkcalendar import *
from tkinter.filedialog import asksaveasfile, asksaveasfilename, askopenfilename




########## window ##########
root = Tk()
root.geometry("1280x720")
root.resizable(0,0)
root.state("zoomed")
root.title("Login")



########## bg image ##########
bg_image = Image.open("bg.jpg")
photo = ImageTk.PhotoImage(bg_image)
bg_panel = Label(root, image=photo)
bg_panel.image = photo
bg_panel.pack(fill="both", expand="yes")

#########search image#################
search_icon = Image.open("search.png")
search_icon = ImageTk.PhotoImage(search_icon)

##########layer###################
layer_icon = Image.open("layer.png")
layer_icon = ImageTk.PhotoImage(layer_icon)


###########upload image###########
# upload_image=ImageTk.PhotoImage(file="show.png")
# upload_image = CTkImage(light_image=Image.open("login.png"),
#                       dark_image=Image.open("login.png"),
#                       size=(150, 150))
upload_img = Image.open("upload.png")
upload_img1 = upload_img.resize((200, 200))
upload_image = ImageTk.PhotoImage(upload_img1)
upload_image1 = ImageTk.PhotoImage(upload_img1)

profile_img = Image.open("profile.png")
profile_img1 = profile_img.resize((40, 40))
profile_img2 = ImageTk.PhotoImage(profile_img1)

# upload_image1 = CTkImage(light_image=Image.open("login.png"),
#                       dark_image=Image.open("login.png"),
#                       size=(150, 150))

###### login image #############
user_image = CTkImage(light_image=Image.open("login.png"),
                      dark_image=Image.open("login.png"),
                      size=(150, 150))

profile_image = CTkImage(light_image=Image.open("login.png"),
                      dark_image=Image.open("login.png"),
                      size=(43, 43))

profile_image1 = CTkImage(light_image=Image.open("profile.png"),
                      dark_image=Image.open("profile.png"),
                      size=(43, 43))

##############logout and delete#############
logout_image = Image.open("logout30.png")
logout_image = ImageTk.PhotoImage(logout_image)

delete_image = Image.open("delete30.png")
delete_image = ImageTk.PhotoImage(delete_image)

######### show and hide icon ################
show_image = Image.open("icon_show.png")
show_icon = ImageTk.PhotoImage(show_image)
# shw_icon = CTkImage(light_image=Image.open("icon_show.png"),
#                       dark_image=Image.open("icon_show.png"),
#                       size=(30, 30))
hide_icon = Image.open("icon_hide.png")
hide_icon = ImageTk.PhotoImage(hide_icon)
# hide_icon = CTkImage(light_image=Image.open("icon_hide.png"),
#                       dark_image=Image.open("icon_hide.png"),
#                       size=(30, 30))
white_theme = Image.open("icons_off.png")
white_theme = ImageTk.PhotoImage(white_theme)

dark_theme = Image.open("icons_on.png")
dark_theme = ImageTk.PhotoImage(dark_theme)

######## menu ############
menu1 = Image.open("menu.png")
menu2 = menu1.resize((40, 40))
menu = ImageTk.PhotoImage(menu2)

close_menu = Image.open("close.png")
close_menu1 = close_menu.resize((40, 40))
close_menu2 = ImageTk.PhotoImage(close_menu1)



############ create database ###############
def create_database():
    if not os.path.exists("accounts_data.db"):
        connection = sqlite3.connect("accounts_data.db")

        cursor = connection.cursor()
        cursor.execute("""
        
        CREATE TABLE accounts(
        
        username text,
        password text
        
        )
        
        """)

        connection.commit()
        connection.close()



def register_account(username, password):
    try:
        connection = sqlite3.connect("accounts_data.db")

        cursor = connection.cursor()
        cursor.execute(f"""

        INSERT INTO accounts VALUES('{username}', '{password}')
                
        """)

        connection.commit()

        cursor.execute("""
        
        SELECT * FROM accounts

        """)


        connection.commit()
        connection.close()
        return True
    except Exception as error:
        return False


def check_username_already_exists(username):
    connection = sqlite3.connect("accounts_data.db")

    cursor = connection.cursor()
    cursor.execute(f"""

    SELECT username FROM accounts WHERE username == '{username}'
               
    """)

    connection.commit()

    response = cursor.fetchall()
    connection.close()
    return(response)



def verify_password(username, password):
    connection = sqlite3.connect("accounts_data.db")

    cursor = connection.cursor()
    cursor.execute(f"""

    SELECT username FROM accounts WHERE username == '{username}'
    AND password == '{password}'
               
    """)

    connection.commit()

    response = cursor.fetchall()
    connection.close()
    return response



def message_box(msg):
    # Check if login_frame exists
    if root.winfo_exists():
        message_frame = Frame(root, relief=SOLID, highlightthickness=2, highlightbackground="gray")
        close_button = Button(message_frame, text="X", font="bold 12", bd=0, command=lambda: message_frame.destroy())
        close_button.pack(side=TOP, anchor=E)
        message_lb = Label(message_frame, text=msg, font="bold 15")
        message_lb.pack(pady=20)

        message_frame.place(x=90, y=270, width=230, height=180)
    else:
        print("Error: login_frame does not exist.")



########## Login frame ############
def login_page():
    global login_frame
    global username
    login_frame=Frame(root,bg="white",width=400,height=600)
    login_frame.place(relx=0.5,rely=0.5,anchor=CENTER)
    login_frame.pack_propagate(False)


######### register forward page function #############
    def register_forward_page():
        login_frame.destroy()
        register_page()



######### functions ###############
    def show_hide():
        def hide():
            show_hide_btn.configure(image=hide_icon)
            show_hide_btn.configure(command=show_hide)
            password.configure(show="●")
            
       
        show_hide_btn.configure(image=show_icon)
        show_hide_btn.configure(command=hide)
        password.configure(show="")
            
                 


############## theme function #########################
    def switch_theme(theme):
        if theme == "Dark":
            root.tk_setPalette("black")

            theme_lb.config(text="Dark")
            theme_btn.config(image=dark_theme)
            theme_btn.config(command= lambda: switch_theme("Light"))

        else:
            root.tk_setPalette("white")

            theme_lb.config(text="Light")
            theme_btn.config(image=white_theme)
            theme_btn.config(command= lambda: switch_theme("Dark"))


    def verify():
        if username.get() != "":
            if password.get() != "":
                if check_username_already_exists(username.get()):
                    if verify_password(username=username.get(),
                                         password=password.get()):
                        messagebox.showinfo("Account", "Successful login")
                        name = username.get()
                        login_frame.destroy()
                        dashboard(username=name)
                                            
                    else:
                         messagebox.showerror("Password", "!Wrong Password")

                else:
                    messagebox.showerror("Username", "!Wrong username")

            else:
                messagebox.showerror("Password", "Password is Required")
        else:
            messagebox.showerror("Username", "Username is Required")
        # if username == username.get() and password == password.get():
        #     login_frame.destroy()
        #     student_database()








    ########### labels############
    login_lb=Label(login_frame,text="Login",font="arial 30 bold",bg="white",fg="#0D47A1")
    login_lb.pack(pady=20)

    username_lb=Label(login_frame,text="Username",font="arial 20",bg="white",fg="#0D47A1")
    username_lb.place(relx=0.5,rely=0.42,anchor=CENTER)

    password_lb=Label(login_frame,text="Password",font="arial 20",bg="white",fg="#0D47A1")
    password_lb.place(relx=0.5,rely=0.63,anchor=CENTER)

    register_lb=Label(login_frame,text="Don't have an account?",font="arial 12",bg="white",fg="#0D47A1")
    register_lb.place(relx=0.35,rely=0.80,anchor=CENTER)

    user_image_lb = CTkLabel(login_frame,text="",
                        image=user_image,
                        anchor="n",
                        compound="top",
                        fg_color="white")
    user_image_lb.place(relx=0.5,rely=0.25,anchor=CENTER)

    theme_lb=Label(login_frame,text="white",font="arial 10",bg="white",fg="black")
    theme_lb.place(relx=0.8,rely=0.86,anchor=CENTER)



    ############  entry ############
    username = CTkEntry(login_frame, placeholder_text="enter username",
                        text_color="black",
                        width=300,bg_color="black",
                        fg_color="white",
                        font=("arial", 15))
    username.place(relx=0.5,rely=0.5,anchor=CENTER)

    password = CTkEntry(login_frame,
                        placeholder_text="enter password",
                        text_color="black",width=300,
                        bg_color="black",
                        fg_color="white",
                        font=("arial", 15),
                        show="●")
    password.place(relx=0.5,rely=0.7,anchor=CENTER)

    ########### button ############
    login_btn = CTkButton(login_frame,text="Login",
                    width=120,
                    height=32,
                    border_width=0,
                    corner_radius=8,
                    fg_color="#0D47A1",
                    command=verify)
    login_btn.place(relx=0.5,rely=0.90,anchor=CENTER)

    register_btn = CTkButton(login_frame,text="Register",
                    width=100,
                    height=20,
                    border_width=0,
                    corner_radius=8,
                    fg_color="#0D47A1",
                    command=register_forward_page)
    register_btn.place(relx=0.70,rely=0.80,anchor=CENTER)

    # show_hide_btn = CTkButton(lgn_frame,image=hide_icon,
    #                          text="",
    #                          bg_color="white",
    #                          fg_color="white",
    #                          width=5,
    #                          command=show_hide_password)
    # show_hide_btn.place(relx=0.94,rely=0.7,anchor=CENTER)

    show_hide_btn = Button(login_frame, text="",
                        bg="white",
                        fg="white",
                        activebackground="white",
                        image=hide_icon,
                        bd=0,
                        command=show_hide)
    show_hide_btn.place(relx=0.94,rely=0.7,anchor=CENTER)

    theme_btn = Button(login_frame, text="",
                        bg="white",
                        fg="white",
                        image=white_theme,
                        bd=0,
                        command=lambda: switch_theme("Dark"))
    theme_btn.place(relx=0.8,rely=0.9,anchor=CENTER)



def register_page():
    register_frame=Frame(root,bg="white",width=400,height=600)
    register_frame.place(relx=0.5,rely=0.5,anchor=CENTER)
    register_frame.pack_propagate(False)


############show hide password############
    def show_hide():
        def hide():
            show_hide_btn.configure(image=hide_icon)
            show_hide_btn.configure(command=show_hide)
            password.configure(show="●")
            confirm_password.configure(show="●")
            
       
        show_hide_btn.configure(image=show_icon)
        show_hide_btn.configure(command=hide)
        password.configure(show="")
        confirm_password.configure(show="")


########### forward login page ##########
    def register_login_page():
        register_frame.destroy()
        login_page()


######## verify function #######################
    def verify():
        if username.get() != "":
            if password.get() != "":
                if confirm_password.get() == "":
                    messagebox.showerror("Password", "Confirm Password")
                    # message_box(msg="Please confirm \nyour password")
                elif confirm_password.get() != password.get():
                    messagebox.showerror("Password", "Confirm Incorrect")
                    # message_box(msg="Password incorrect")
                elif not check_username_already_exists(username.get()):
                    response = register_account(username=username.get(),
                                                password=password.get()) 
                    if response:
                        username.delete(0, END)
                        password.delete(0, END)
                        confirm_password.delete(0, END)
                        messagebox.showinfo("Account", "Account created successfully")
                else:
                    messagebox.showerror("username", "Username Already exists")
            else:
                messagebox.showerror("Password", "Password is Required")        
        else:
            messagebox.showerror("Username","Username is Required")



########### labels############
    create_account_lb=Label(register_frame,text="Create Account",font="arial 30 bold",bg="white",fg="#0D47A1")
    create_account_lb.pack(pady=20)

    username_lb=Label(register_frame,text="Username",font="arial 20",bg="white",fg="#0D47A1")
    username_lb.place(relx=0.5,rely=0.21,anchor=CENTER)

    password_lb=Label(register_frame,text="Password",font="arial 20",bg="white",fg="#0D47A1")
    password_lb.place(relx=0.5,rely=0.42,anchor=CENTER)

    Cpassword_lb=Label(register_frame,text="Confirm Password",font="arial 20",bg="white",fg="#0D47A1")
    Cpassword_lb.place(relx=0.5,rely=0.63,anchor=CENTER)

############  entry ############

    username = CTkEntry(register_frame, 
                          placeholder_text="Username",
                          text_color="black",
                          width=300,
                          bg_color="black",
                          fg_color="white",
                          font=("arial", 15))
    username.place(relx=0.5,rely=0.3,anchor=CENTER)
    password = CTkEntry(register_frame, 
                          placeholder_text="Password",
                          text_color="black",
                          width=300,
                          bg_color="black",
                          fg_color="white",
                          show="●")
    password.place(relx=0.5,rely=0.5,anchor=CENTER)

    confirm_password = CTkEntry(register_frame, 
                        placeholder_text="Confirm Password",
                        text_color="black",width=300,
                        bg_color="black",
                        fg_color="white",
                        show="●")
    confirm_password.place(relx=0.5,rely=0.7,anchor=CENTER)


########### button ############
    login_btn = CTkButton(register_frame,text="Login",
                    width=120,
                    height=32,
                    border_width=0,
                    corner_radius=8,
                    fg_color="#0D47A1",
                    command=register_login_page)
    login_btn.place(relx=0.5,rely=0.90,anchor=CENTER)

    register_btn = CTkButton(register_frame,text="Register",
                    width=150,
                    height=32,
                    border_width=0,
                    corner_radius=8,
                    fg_color="#0D47A1",
                    command=verify)
    register_btn.place(relx=0.5,rely=0.80,anchor=CENTER)

    show_hide_btn = Button(register_frame, text="",
                        bg="white",
                        fg="white",
                        image=hide_icon,
                        bd=0,
                        command=show_hide)
    show_hide_btn.place(relx=0.94,rely=0.7,anchor=CENTER)

color1="#666666"
color2="#663300"
color3="#996666"

# def about():
#     dashboard_frame.destroy()
#     menu_frame.destroy()
#     menu_btn.configure(image=menu)
#     menu_btn.configure(command=show_menu)
#     about_us()

# def show_dashboard():
#     about_frame.destroy()
#     menu_frame.destroy()
#     menu_btn.configure(image=menu)
#     menu_btn.configure(command=show_menu)
#     dashboard(username)


# def about_us():
#     global about_frame
#     about_frame=Frame(root,bg="white",width=1280,height=720,background=color1)
#     about_frame.place(relx=0.5,rely=0.5,anchor=CENTER)
#     about_frame.pack_propagate(False)


# def show_menu():
#     global menu_frame
#     def hide_menu():
#         menu_btn.configure(image=menu)
#         menu_btn.configure(command=show_menu)
#         menu_frame.destroy()
   
#     menu_btn.configure(image=close_menu2)
#     menu_btn.configure(command=hide_menu)
#     menu_frame=Frame(root,width=300,height=720,bg=color2)
#     menu_frame.place(x=43,y=10)
#     dash_btn=Button(menu_frame,text="Dashboard",font="arial 18 bold",fg="white",bg=color2,
#                     bd=0,command=show_dashboard)
#     dash_btn.place(x=80,y=50)
#     about_btn=Button(menu_frame,text="About Us",font="arial 18 bold",fg="white",
#                         bg=color2,bd=0,command=about)
#     about_btn.place(x=80,y=600)

        

# ########### menu button ############
# menu_btn = Button(root,text="",image=menu,bg=color1,bd=0,command=show_menu)
# menu_btn.place(x=0,y=12)

def dashboard(username):
    global dashboard
    global dashboard_frame
    dashboard_frame=Frame(root,bg="white",width=1280,height=720,background=color1)
    dashboard_frame.place(relx=0.5,rely=0.5,anchor=CENTER)
    dashboard_frame.pack_propagate(False)

    
    def verify():
        sure = askyesno(title='Delete Account',
                        message='Are you Sure you want to delete the account')
        if sure:
            delete_account(username)
            logout()
            messagebox.showinfo("Account", 'Account deleted successfully')


    def delete_account(username):
        connection = sqlite3.connect("accounts_data.db")

        cursor = connection.cursor()
        cursor.execute(f"""

        DELETE FROM accounts WHERE username == '{username}'
                
        """)

        connection.commit()

        response = cursor.fetchall()
        connection.close()

    def logout():
        dashboard_frame.destroy()
        login_page()

    def profile():
        def hide_profile():
            logout_btn.destroy()
            delete_account_btn.destroy()
            profile_btn.configure(command=profile)
        logout_btn=Button(dashboard_frame,text="",
                          image=logout_image,
                            bg="white",
                            fg="white",
                            font="arial 12",
                            width=30,
                            height=30,
                            bd=0,
                            command=logout)
        logout_btn.place(relx=0.97,rely=0.20,anchor=CENTER)

        delete_account_btn=Button(dashboard_frame,
                              text="",
                              image=delete_image,
                              bg="white",
                              fg="white",
                              font="arial 12",
                              width=30,
                              height=30,
                              bd=0,
                              command=verify)
        delete_account_btn.place(relx=0.97,rely=0.27,anchor=CENTER)
        profile_btn.configure(command=hide_profile)

    def selection():
        global gender
        value=radio.get()
        if value == 1:
            gender="Male"
        else:
            gender="Female"
            
    # def show_image():
    #     filename=filedialog.askopenfilename(initialdir=os.getcwd(),
    #                                         title="Select image file",
    #                                         filetypes=(('JPG', '*.jpg'),
    #                                                    ('PNG', '*.png'),
    #                                                    ('All files', '*.txt')))
    #     upload_image = (Image.open(filename))
    #     resized_image=upload_image.resize((190,190), Image.LANCZOS)
    #     photo2 = ImageTk.PhotoImage(resized_image)
    #     image_lb.configure(image=photo2)
    #     image_lb.image=photo2 
        

    ###########date selection###########
    def pick_date(grab_date):
        # global date_window
        # global ca1
        def grab_date():
            date_entry.delete(0, END)
            date_entry.insert(0, ca1.get_date())
            date_window.destroy()

        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+590+370')
        ca1 = Calendar(date_window, selectmode="day", date_pattern="mm/dd/yyyy")
        ca1.place(x=0,y=0)

        submit_btn=Button(date_window, text="Submit",command=grab_date)
        submit_btn.place(x=80,y=190)

    def pick_date1(grab_date1):
        # global date_window
        # global ca1
        def grab_date1():
            dob_entry.delete(0, END)
            dob_entry.insert(0, ca1.get_date())
            date_window.destroy()

        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+590+370')
        ca1 = Calendar(date_window, selectmode="day", date_pattern="mm/dd/yyyy")
        ca1.place(x=0,y=0)

        submit_btn=Button(date_window, text="Submit",command=grab_date1)
        submit_btn.place(x=80,y=190)

    ############define registration no############
    def registration_no():
        file=openpyxl.load_workbook('student_data.xlsx')
        sheet=file.active
        row=sheet.max_row

        max_row_value=sheet.cell(row=row,column=1).value
        try:
            Registration.set(max_row_value+1)
        except:
            Registration.set("1")

    
    #########clear#############
    def clear():
        # global upload_image
        # global upload_image1
        Name.set('')
        DOB.set('')
        Religion.set('')
        Skills.set('')
        Father.set('')
        Father_Occupation.set('')
        Mother.set('')
        Mother_Occupation.set('')
        Class.set("Select Class")

        registration_no()

        

        save_btn.configure(state='normal')

        
        # image_lb.configure(image=upload_image1)
        # image_lb.image=upload_image1
        
        # upload_image=""
        

    def save():
        R1=Registration.get()
        N1=Name.get()
        C1=Class.get() 
        try:
            G1=gender
        except:
            messagebox.showerror("Error", "Select Gender!")

        D2=DOB.get()
        D1=Date.get()
        R2=Religion.get()
        S1=Skills.get()
        F1=Father.get()
        F2=Father_Occupation.get()
        M1=Mother.get()
        M2=Mother_Occupation.get()

        if N1 == "" or C1 == "Select Class" or D2 == ""or D1 =="" or R2 == "" or S1 == "" or F1 == "" or F2 == "" or  M1 == "" or M2 == "":
            messagebox.showerror("Error", "Few data missing")
        else:
            file=openpyxl.load_workbook('student_data.xlsx')
            sheet=file.active
            sheet.cell(column=1,row=sheet.max_row+1,value=R1)
            sheet.cell(column=2,row=sheet.max_row,value=N1)
            sheet.cell(column=3,row=sheet.max_row,value=C1)
            sheet.cell(column=4,row=sheet.max_row,value=G1)
            sheet.cell(column=5,row=sheet.max_row,value=D2)
            sheet.cell(column=6,row=sheet.max_row,value=D1)
            sheet.cell(column=7,row=sheet.max_row,value=R2)
            sheet.cell(column=8,row=sheet.max_row,value=S1)
            sheet.cell(column=9,row=sheet.max_row,value=F1)
            sheet.cell(column=10,row=sheet.max_row,value=M1)
            sheet.cell(column=11,row=sheet.max_row,value=F2)
            sheet.cell(column=12,row=sheet.max_row,value=M2)

            file.save(r'student_data.xlsx')

            # try:
            #     pass
            #     # image=filedialog.asksaveasfilename(filetypes=[('JPG', '*.jpg'),
            #     #                                        ('PNG', '*.png'),
            #     #                                        ('All files', '*.txt')],
            #     #                                        defaultextension='.jpg')
            #     #                                         #('student images/'+str(R1)+'.jpg')
            #     # imageob=open(image,'w')
            #     # imageob.write("")
            #     # imageob.close()
            #     # upload_image.save('student images/'+str(R1)+'.jpg')
            # except:
            #     messagebox.showerror("Info", "Profile Picture is not available")
            messagebox.showinfo("Info", "Data Entered Successfully!")

            clear()
            registration_no()
               

    def search():
        text = search_entry.get()
        clear()
        save_btn.configure(state='disable')

        file = openpyxl.load_workbook('student_data.xlsx')
        sheet = file.active

        found = False  # Flag to check if a match is found

        for row in sheet.rows:
            reg_number = row[0].value
            student_name = row[1].value  # Assuming the name is in the second column

            # Check if the search text matches the registration number
            if reg_number and str(reg_number) == text:
                found = True
                name = row[0]
                registration_no_position = str(name)[14:-1]
                registration_number = str(name)[15:-1]
                break

            # Check if the search text matches the student's name
            if student_name and text.lower() in student_name.lower():
                found = True
                name = row[0]
                registration_no_position = str(name)[14:-1]
                registration_number = str(name)[15:-1]
                break

        if not found:
            messagebox.showerror("Invalid", "Invalid registration number or name!")
            return

        try:
            registration_no_position = str(name)[14:-1]
        except:
            messagebox.showerror("Invalid", "Invalid registration number!")

        x1 = sheet.cell(row=int(registration_number), column=1).value
        x2 = sheet.cell(row=int(registration_number), column=2).value
        x3 = sheet.cell(row=int(registration_number), column=3).value
        x4 = sheet.cell(row=int(registration_number), column=4).value
        x5 = sheet.cell(row=int(registration_number), column=5).value
        x6 = sheet.cell(row=int(registration_number), column=6).value
        x7 = sheet.cell(row=int(registration_number), column=7).value
        x8 = sheet.cell(row=int(registration_number), column=8).value
        x9 = sheet.cell(row=int(registration_number), column=9).value
        x10 = sheet.cell(row=int(registration_number), column=10).value
        x11 = sheet.cell(row=int(registration_number), column=11).value
        x12 = sheet.cell(row=int(registration_number), column=12).value

        Registration.set(x1)
        Name.set(x2)
        Class.set(x3)

        if x4 == 'Female':
            R2.select()
        else:
            R1.select()

        DOB.set(x5)
        Date.set(x6)
        Religion.set(x7)
        Skills.set(x8)
        Father.set(x9)
        Mother.set(x10)
        Father_Occupation.set(x11)
        Mother_Occupation.set(x12)

        # upload_image = (Image.open('student images/'+str(x1)+'.jpg'))
        # resized_image=upload_image.resize((190,190), Image.LANCZOS)
        # photo2 = ImageTk.PhotoImage(resized_image)
        # image_lb.configure(image=photo2)
        # image_lb.image=photo2 


    def update():
        R1=Registration.get()
        N1=Name.get()
        C1=Class.get() 
        selection()
        G1=gender
        D2=DOB.get()
        D1=Date.get()
        R2=Religion.get()
        S1=Skills.get()
        F1=Father.get()
        F2=Father_Occupation.get()
        M1=Mother.get()
        M2=Mother_Occupation.get()

        file=openpyxl.load_workbook('student_data.xlsx')
        sheet=file.active

        for row in sheet.rows:
            if row[0].value == R1:
                name=row[0]
                registration_no_position=str(name)[14:-1]
                registration_number=str(name)[15:-1]
        #sheet.cell(column=1,row=int(registration_no),value=R1)
        sheet.cell(column=2,row=int(registration_number),value=N1)
        sheet.cell(column=3,row=int(registration_number),value=C1)
        sheet.cell(column=4,row=int(registration_number),value=G1)
        sheet.cell(column=5,row=int(registration_number),value=D2)
        sheet.cell(column=6,row=int(registration_number),value=D1)
        sheet.cell(column=7,row=int(registration_number),value=R2)
        sheet.cell(column=8,row=int(registration_number),value=S1)
        sheet.cell(column=9,row=int(registration_number),value=F1)
        sheet.cell(column=10,row=int(registration_number),value=M1)
        sheet.cell(column=11,row=int(registration_number),value=F2)
        sheet.cell(column=12,row=int(registration_number),value=M2)

        file.save(r'student_data.xlsx')

        # try:
        #     upload_image.save('student images/'+str(R1)+'.jpg')
        # except:
        #     messagebox.showerror("Info", "Profile Picture is not available")
        messagebox.showinfo("Update", "Update Successfully")
        clear()

    def delete():
        text = search_entry.get()  # Assuming this is used to identify the student
        file = openpyxl.load_workbook('student_data.xlsx')
        sheet = file.active

        found = False  # Flag to check if a match is found

        for row in sheet.iter_rows(min_row=2):  # Assuming the first row is headers
            reg_number = row[0].value

            # Check if the search text matches the registration number
            if reg_number and str(reg_number) == text:
                found = True
                sheet.delete_rows(row[0].row, 1)  # Delete the entire row
                file.save('student_data.xlsx')  # Save changes to the file
                clear()  # Clear all input fields
                messagebox.showinfo("Deleted", "Student record deleted successfully.")
                break

        if not found:
            messagebox.showerror("Not Found", "Student record not found.")


#############top label######################
    mail_lb=Label(dashboard_frame,text="Email: kalcomputerstech@gmail.com",width=10,height=3,bg=color3,anchor='e')
    mail_lb.pack(side=TOP,fill=X)
    registration_lb=Label(dashboard_frame,text="STUDENT REGISTRATION",width=10,height=2,bg=color2,
                          fg="white",font="arial 20 bold",
                          highlightbackground="white",
                          highlightthickness=1)
    registration_lb.pack(side=TOP,fill=X)
    dashboard_lb=Label(dashboard_frame,text="Dashboard",width=10,height=2,bg=color2,fg="white",font="arial 18 bold")
    dashboard_lb.place(x=50,y=55)


#############profile button#########
    profile_btn = CTkButton(dashboard_frame,text="",
                    fg_color=color2,
                    bg_color=color2,
                    image=profile_image1,
                    command=profile,
                    width=5,height=5)
    profile_btn.place(relx=0.975,rely=0.12,anchor=CENTER)
    ############ Create entry database ###########

    file = pathlib.Path('student_data.xlsx')
    if file.exists():
        pass
    else:
        file=Workbook()
        sheet=file.active
        sheet['A1']="Registration No."
        sheet['B1']="Name"
        sheet['C1']="Class"
        sheet['D1']="Gender"
        sheet['E1']="DOB"
        sheet['F1']="Date of Registration"
        sheet['G1']="Religion"
        sheet['H1']="Skill"
        sheet['I1']="Father's Name"
        sheet['J1']="Mother's Name"
        sheet['K1']="Father's Occupation"
        sheet['L1']="Mother's Occupation"
        file.save(r'student_data.xlsx')

    #######search box##################
    Search = StringVar()
    search_entry=Entry(dashboard_frame,textvariable=Search,width=15,bd=2,font="arial 20")
    search_entry.place(x=820,y=70)
   
    search_btn=Button(dashboard_frame,text="Search",compound=LEFT,image=search_icon,width=123,
                      height=43,bg="#68ddfa",font="arial 13 bold",command=search)
    search_btn.place(x=1060,y=66)

    ######################layer label############

    layer_lb=Label(dashboard_frame,image=layer_icon,bg=color2,bd=1)
    layer_lb.place(x=10,y=60)

    ###############Registration and Date#################
    Label(dashboard_frame,text="Registration No:",font="arial 13",bg=color1,fg="#ededed").place(x=30,y=150)
    Label(dashboard_frame,text="Date:",font="arial 13",bg=color1,fg="#ededed").place(x=500,y=150)

    Registration=IntVar()
    Date=StringVar()

    reg_entry=Entry(dashboard_frame,textvariable=Registration,width=15,font="arial 10")
    reg_entry.place(x=160,y=150)

    date_entry=Entry(dashboard_frame,textvariable=Date,width=15,font="arial 10")
    date_entry.place(x=550,y=150)
    date_entry.insert(0, "dd/mm/yyyy")
    date_entry.bind("<1>",pick_date)

    registration_no() ########Registration no.############


    ###################Student Details#########################
    frame=LabelFrame(dashboard_frame,text="Student's Details",font=20,bd=2,width=900,bg="#ededed",fg="black",relief=GROOVE,height=250)
    frame.place(x=30,y=200)

    Label(frame,text="Full Name:",font="arial 13",bg="#ededed",fg="black").place(x=30,y=50)
    Label(frame,text="Date of Birth:",font="arial 13",bg="#ededed",fg="black").place(x=30,y=100)
    Label(frame,text="Gender:",font="arial 13",bg="#ededed",fg="black").place(x=30,y=150)

    Label(frame,text="Class:",font="arial 13",bg="#ededed",fg="black").place(x=500,y=50)
    Label(frame,text="Religion:",font="arial 13",bg="#ededed",fg="black").place(x=500,y=100)
    Label(frame,text="Skills:",font="arial 13",bg="#ededed",fg="black").place(x=500,y=150)

    Name=StringVar()
    name_entry=Entry(frame,textvariable=Name,width=20,font="arial 10")
    name_entry.place(x=160,y=50)

    DOB=StringVar()
    dob_entry=Entry(frame,textvariable=DOB,width=20,font="arial 10")
    dob_entry.place(x=160,y=100)
    dob_entry.insert(0, "dd/mm/yyyy")
    dob_entry.bind("<1>",pick_date1)

    radio=IntVar()
    R1=Radiobutton(frame,text="Male",variable=radio,value=1,bg="#ededed",fg="black",command=selection)
    R1.place(x=150,y=150)

    R2=Radiobutton(frame,text="Female",variable=radio,value=2,bg="#ededed",fg="black",command=selection)
    R2.place(x=200,y=150)

    Religion=StringVar()
    religion_entry=Entry(frame,textvariable=Religion,width=20,font="arial 10")
    religion_entry.place(x=630,y=100)

    Skills=StringVar()
    skills_entry=Entry(frame,textvariable=Skills,width=20,font="arial 10")
    skills_entry.place(x=630,y=150)


    Class=ttk.Combobox(frame,values=['NURSERY 1','NURSERY 2','KG 1','KG 2','PRIMARY 1',
                                     'PRIMARY 2','PRIMARY 3','PRIMARY 4','PRIMARY 5',
                                     'PRIMARY 6','JHS 1','JHS 2','JHS 3'],
                                     font="Roboto 10",width=17,state="r")
    Class.place(x=630,y=50)
    Class.set("Select Class")

    

    ###################Parents Details#########################
    frame1=LabelFrame(dashboard_frame,text="Parent's Details",font=20,bd=2,width=900,bg="#ededed",fg="black",relief=GROOVE,height=220)
    frame1.place(x=30,y=470)
   
    Label(frame1,text="Father's Name:",font="arial 13",bg="#ededed",fg="black").place(x=30,y=50)

    Label(frame1,text="Occupation:",font="arial 13",bg="#ededed",fg="black").place(x=30,y=100)

    Father=StringVar()
    father_entry=Entry(frame1,textvariable=Father,width=20,font="arial 10")
    father_entry.place(x=160,y=50)

    Father_Occupation=StringVar()
    fo_entry=Entry(frame1,textvariable=Father_Occupation,width=20,font="arial 10")
    fo_entry.place(x=160,y=100)

    Label(frame1,text="Mother's Name:",font="arial 13",bg="#ededed",fg="black").place(x=500,y=50)

    Label(frame1,text="Occupation:",font="arial 13",bg="#ededed",fg="black").place(x=500,y=100)

    Mother=StringVar()
    mother_entry=Entry(frame1,textvariable=Mother,width=20,font="arial 10")
    mother_entry.place(x=630,y=50)

    Mother_Occupation=StringVar()
    mo_entry=Entry(frame1,textvariable=Mother_Occupation,width=20,font="arial 10")
    mo_entry.place(x=630,y=100)

      
    # frame2=Frame(dashboard,bd=3,bg="black",width=200,height=200,relief=GROOVE)
    # frame2.place(x=1000,y=150)
    # image_lb=Label(frame2,text="",bg="white",image=upload_image,width=200,height=200)
    # image_lb.place(x=0,y=0)

    ###########buttons#############
    
    save_btn=Button(dashboard_frame,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=save)
    save_btn.place(x=1000,y=290)

    reset_btn=Button(dashboard_frame,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightblue",command=clear)
    reset_btn.place(x=1000,y=370)

    update_btn=Button(dashboard_frame,text="Update",width=19,height=2,font="arial 12 bold",bg="lightpink",command=update)
    update_btn.place(x=1000,y=450)

    delete_btn=Button(dashboard_frame,text="Delete",width=19,height=2,font="arial 12 bold",bg="red",command=delete)
    delete_btn.place(x=1000,y=530)

   
####### run ###########
create_database()
login_page()
root.mainloop()